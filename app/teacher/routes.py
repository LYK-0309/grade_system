"""
教师路由: 仪表板/成绩录入/成绩修改/查询/导出
"""
from flask import (render_template, redirect, url_for, request, flash,
                   jsonify, send_file, current_app, abort)
from flask_login import login_required, current_user
from app.models import (db, User, ClassInfo, Student, Teacher, Course, Grade,
                        GradeRevision, Notification)
from app.utils import (role_required, teacher_or_admin, log_operation,
                       create_grade_export, export_to_bytes, calc_class_stats,
                       calc_grade_distribution, get_rankings)
from datetime import datetime
import io
import openpyxl

from app.teacher import teacher_bp


@teacher_bp.route('/dashboard')
@login_required
@teacher_or_admin
def dashboard():
    """教师仪表板"""
    teacher = current_user.teacher
    if teacher:
        courses = teacher.courses.order_by(Course.semester.desc()).all()
    else:
        courses = Course.query.order_by(Course.semester.desc()).limit(10).all()
    stats = {
        'course_count': len(courses),
        'student_count': sum(c.class_info.students.count() if c.class_info else 0 for c in courses),
        'grade_count': sum(c.grades.count() for c in courses),
        'pending_revisions': GradeRevision.query.filter_by(status='pending').count(),
    }
    return render_template('teacher/dashboard.html', courses=courses, stats=stats)


# ── 成绩录入 ──────────────────────────────────────────
@teacher_bp.route('/grades/entry/<int:course_id>', methods=['GET', 'POST'])
@login_required
@teacher_or_admin
def grade_entry(course_id):
    course = Course.query.get_or_404(course_id)
    if course.is_locked and not current_user.is_super_admin:
        flash('该课程成绩已锁定，如需修改请申请修改', 'warning')
        return redirect(url_for('teacher.course_detail', course_id=course_id))

    if not course.class_info:
        flash('该课程未绑定班级', 'warning')
        return redirect(url_for('teacher.dashboard'))

    students = course.class_info.students.filter_by(status='正常').order_by(Student.student_no).all()

    if request.method == 'POST':
        exam_type = request.form.get('exam_type', '期末')
        updated = 0
        for stu in students:
            score_raw = request.form.get(f'score_{stu.id}', '').strip()
            special = request.form.get(f'special_{stu.id}', '')
            score_discuss = request.form.get(f'discuss_{stu.id}', '', type=float)
            score_homework = request.form.get(f'homework_{stu.id}', '', type=float)
            score_attendance = request.form.get(f'attendance_{stu.id}', '', type=float)
            score_points = request.form.get(f'points_{stu.id}', '', type=float)
            score_pbl = request.form.get(f'pbl_{stu.id}', '', type=float)

            grade = Grade.query.filter_by(
                student_id=stu.id, course_id=course.id, exam_type=exam_type
            ).first()

            if not score_raw and not special:
                continue

            score = None
            if special in ('缺考', '作弊', '缓考'):
                score = None
            else:
                try:
                    score = float(score_raw) if score_raw else None
                    if score is not None and (score < 0 or score > 100):
                        flash(f'{stu.name} 的分数超出范围(0-100)', 'danger')
                        return render_template('teacher/grade_entry.html',
                                               course=course, students=students)
                except ValueError:
                    continue

            if grade is None:
                grade = Grade(student_id=stu.id, course_id=course.id, exam_type=exam_type,
                              created_by=current_user.id)
                db.session.add(grade)

            grade.score = score
            grade.special_mark = special or None
            grade.score_discuss = score_discuss
            grade.score_homework = score_homework
            grade.score_attendance = score_attendance
            grade.score_points = score_points
            grade.score_pbl = score_pbl
            grade.gpa = Grade.calc_gpa(score, special)
            grade.updated_at = datetime.now()
            updated += 1

        db.session.commit()
        log_operation('create', '成绩录入', f'课程 {course.name}, 共 {updated} 条')
        flash(f'成功录入 {updated} 条成绩', 'success')
        return redirect(url_for('teacher.course_detail', course_id=course_id))

    # 查看已有成绩
    exam_type = request.args.get('exam_type', '期末')
    existing = {}
    for g in Grade.query.filter_by(course_id=course.id, exam_type=exam_type).all():
        existing[g.student_id] = g

    return render_template('teacher/grade_entry.html',
                           course=course, students=students, existing=existing, exam_type=exam_type)


@teacher_bp.route('/grades/import/<int:course_id>', methods=['POST'])
@login_required
@teacher_or_admin
def grade_import(course_id):
    """Excel批量导入成绩"""
    course = Course.query.get_or_404(course_id)
    if course.is_locked:
        flash('课程已锁定', 'danger')
        return redirect(url_for('teacher.course_detail', course_id=course_id))
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('请上传Excel文件', 'danger')
        return redirect(url_for('teacher.grade_entry', course_id=course_id))
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        imported = 0
        exam_type = request.form.get('exam_type', '期末')
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            student_no = str(row[0]).strip()
            stu = Student.query.filter_by(student_no=student_no).first()
            if not stu:
                continue
            score = None
            special = ''
            if isinstance(row[1], (int, float)):
                score = float(row[1])
            elif row[1] and str(row[1]).strip() in ('缺考', '作弊', '缓考'):
                special = str(row[1]).strip()

            grade = Grade.query.filter_by(student_id=stu.id, course_id=course.id, exam_type=exam_type).first()
            if not grade:
                grade = Grade(student_id=stu.id, course_id=course.id, exam_type=exam_type,
                              created_by=current_user.id)
                db.session.add(grade)
            grade.score = score
            grade.special_mark = special or None
            grade.gpa = Grade.calc_gpa(score, special)
            imported += 1
        db.session.commit()
        log_operation('import', '导入成绩', f'课程 {course.name}, {imported} 条')
        flash(f'成功导入 {imported} 条成绩', 'success')
    except Exception as e:
        flash(f'导入失败: {str(e)}', 'danger')
    return redirect(url_for('teacher.course_detail', course_id=course_id))


@teacher_bp.route('/grades/template/<int:course_id>')
@login_required
@teacher_or_admin
def grade_template(course_id):
    """下载成绩录入模板"""
    course = Course.query.get_or_404(course_id)
    if not course.class_info:
        abort(404)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成绩录入模板'
    ws.append(['学号', '姓名', '成绩', '特殊标记(缺考/作弊/缓考)'])
    for stu in course.class_info.students.filter_by(status='正常').order_by(Student.student_no).all():
        ws.append([stu.student_no, stu.name, '', ''])
    for col, w in enumerate([20, 15, 10, 25], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'{course.name}_成绩录入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 成绩修改申请 ──────────────────────────────────────────
@teacher_bp.route('/grades/<int:grade_id>/revise', methods=['POST'])
@login_required
@teacher_or_admin
def grade_revise(grade_id):
    grade = Grade.query.get_or_404(grade_id)
    if grade.is_locked and not current_user.is_super_admin:
        flash('成绩已锁定', 'danger')
        return redirect(url_for('teacher.course_detail', course_id=grade.course_id))
    new_score = request.form.get('new_score', type=float)
    new_special = request.form.get('new_special', '')
    reason = request.form.get('reason', '')

    rev = GradeRevision(
        grade_id=grade.id,
        original_score=grade.score,
        new_score=new_score if not new_special else None,
        original_special=grade.special_mark,
        new_special=new_special or None,
        modified_by=current_user.id,
        modifier_name=current_user.name,
        reason=reason,
        status='pending'
    )
    db.session.add(rev)
    db.session.commit()
    log_operation('create', '成绩修改申请', f'{grade.student.name} {grade.course.name}: {grade.score} -> {new_score}')
    flash('修改申请已提交，等待教务审核', 'success')
    return redirect(url_for('teacher.course_detail', course_id=grade.course_id))


# ── 课程详情 & 成绩查询 ──────────────────────────────────────────
@teacher_bp.route('/courses/<int:course_id>')
@login_required
@teacher_or_admin
def course_detail(course_id):
    course = Course.query.get_or_404(course_id)
    exam_type = request.args.get('exam_type', '期末')
    grades = Grade.query.filter_by(course_id=course.id, exam_type=exam_type).all()

    # 关联学生
    grade_map = {g.student_id: g for g in grades}
    students = []
    if course.class_info:
        students = course.class_info.students.order_by(Student.student_no).all()
        for stu in students:
            stu._grade = grade_map.get(stu.id)

    stats = calc_class_stats(grades)
    distribution = calc_grade_distribution(grades)
    rankings = get_rankings(grades)

    return render_template('teacher/course_detail.html',
                           course=course, students=students, grades=grades,
                           stats=stats, distribution=distribution, rankings=rankings,
                           exam_type=exam_type, exam_types=['期中', '期末', '补考', '重修'])


# ── 成绩查询 ──────────────────────────────────────────
@teacher_bp.route('/grades/search')
@login_required
@teacher_or_admin
def grade_search():
    class_id = request.args.get('class_id', type=int)
    course_id = request.args.get('course_id', type=int)
    keyword = request.args.get('keyword', '')
    exam_type = request.args.get('exam_type', '')

    query = Grade.query.join(Student).join(Course)
    if class_id:
        query = query.filter(Student.class_id == class_id)
    if course_id:
        query = query.filter(Grade.course_id == course_id)
    if exam_type:
        query = query.filter(Grade.exam_type == exam_type)
    if keyword:
        query = query.filter(Student.name.contains(keyword) | Student.student_no.contains(keyword))

    grades = query.order_by(Grade.course_id, Student.student_no).all()
    classes = ClassInfo.query.all()
    courses = Course.query.all()

    return render_template('teacher/grade_search.html', grades=grades,
                           classes=classes, courses=courses,
                           class_id=class_id, course_id=course_id,
                           keyword=keyword, exam_type=exam_type)


# ── 导出 ──────────────────────────────────────────
@teacher_bp.route('/grades/export/<int:course_id>')
@login_required
@teacher_or_admin
def grade_export(course_id):
    course = Course.query.get_or_404(course_id)
    exam_type = request.args.get('exam_type', '期末')
    grades = Grade.query.filter_by(course_id=course.id, exam_type=exam_type).order_by(
        Grade.score.desc() if True else Student.student_no
    ).all()

    headers = ['排名', '学号', '姓名', '班级', '讨论(20%)', '作业(30%)', '签到(10%)',
               '课程积分(20%)', 'PBL(20%)', '综合成绩', '等级', '绩点', '特殊标记']
    rows = []
    rankings = get_rankings(grades)
    for g in grades:
        stu = g.student
        rows.append([
            rankings.get(g.student_id, '-'),
            stu.student_no, stu.name, stu.class_info.name if stu.class_info else '',
            g.score_discuss or '', g.score_homework or '', g.score_attendance or '',
            g.score_points or '', g.score_pbl or '',
            g.score_display, g.grade_level, f'{g.gpa:.1f}' if g.gpa else '0.0',
            g.special_mark or ''
        ])

    stats = calc_class_stats(grades)
    summary_row = ['', '', f'平均分: {stats["avg"]}', '', '', '', '', '', '',
                   f'及格率: {stats["pass_rate"]}%', '', f'平均GPA: {stats["gpa_avg"]}', '']

    title = f'{course.name} - {course.class_info.name if course.class_info else ""} - {exam_type}成绩'
    wb = create_grade_export(title, headers, rows, summary_row)
    buf = export_to_bytes(wb)
    log_operation('export', '导出成绩', f'课程 {course.name}, {len(rows)} 条')
    return send_file(buf, as_attachment=True,
                     download_name=f'{course.name}_{exam_type}成绩.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 补考管理 ──────────────────────────────────────────
@teacher_bp.route('/makeup/<int:course_id>')
@login_required
@teacher_or_admin
def makeup_list(course_id):
    course = Course.query.get_or_404(course_id)
    # 查找期末不及格的学生
    fail_grades = Grade.query.filter_by(
        course_id=course.id, exam_type='期末'
    ).filter(
        (Grade.score < 60) | (Grade.special_mark.in_(['缺考', '作弊']))
    ).all()

    # 检查是否已有补考成绩
    for g in fail_grades:
        g._makeup = Grade.query.filter_by(
            student_id=g.student_id, course_id=course.id, exam_type='补考'
        ).first()

    return render_template('teacher/makeup_list.html', course=course, fail_grades=fail_grades)


@teacher_bp.route('/makeup/<int:course_id>/entry', methods=['POST'])
@login_required
@teacher_or_admin
def makeup_entry(course_id):
    course = Course.query.get_or_404(course_id)
    student_id = request.form.get('student_id', type=int)
    score = request.form.get('score', type=float)
    special = request.form.get('special', '')

    grade = Grade.query.filter_by(
        student_id=student_id, course_id=course.id, exam_type='补考'
    ).first()
    if not grade:
        grade = Grade(student_id=student_id, course_id=course.id, exam_type='补考',
                      created_by=current_user.id)
        db.session.add(grade)
    grade.score = None if special else score
    grade.special_mark = special or None
    grade.gpa = Grade.calc_gpa(score if not special else None, special)
    db.session.commit()
    log_operation('create', '补考成绩录入', f'{grade.student.name} {course.name}')
    flash('补考成绩已录入', 'success')
    return redirect(url_for('teacher.makeup_list', course_id=course_id))
