"""
管理员路由: 用户管理/基础信息管理/系统设置/日志/通知/备份
"""
from flask import render_template, redirect, url_for, request, flash, jsonify, send_file, current_app
from flask_login import login_required, current_user
from app.models import (db, User, ClassInfo, Student, Teacher, Course, Grade,
                        LoginLog, OperationLog, Notification, SystemSetting, GradeRevision)
from app.utils import (admin_required, role_required, log_operation,
                       create_grade_export, export_to_bytes)
from datetime import datetime, date
import io
import os
import shutil
import openpyxl

from app.admin import admin_bp


# ── 首页 ──────────────────────────────────────────
@admin_bp.route('/')
@login_required
def index():
    """系统首页 - 根据角色重定向"""
    if current_user.is_super_admin:
        return redirect(url_for('admin.dashboard'))
    elif current_user.is_teacher:
        return redirect(url_for('teacher.dashboard'))
    elif current_user.is_student:
        return redirect(url_for('student.dashboard'))
    elif current_user.is_academic:
        return redirect(url_for('admin.academic_dashboard'))
    return redirect(url_for('auth.login'))


@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    """超管仪表板"""
    stats = {
        'user_count': User.query.count(),
        'student_count': Student.query.count(),
        'teacher_count': Teacher.query.count(),
        'class_count': ClassInfo.query.count(),
        'course_count': Course.query.count(),
        'grade_count': Grade.query.count(),
        'teacher_users': User.query.filter_by(role='teacher').count(),
        'student_users': User.query.filter_by(role='student').count(),
        'active_users': User.query.filter_by(is_active_user=True).count(),
    }
    recent_logs = OperationLog.query.order_by(OperationLog.created_at.desc()).limit(10).all()
    return render_template('admin/dashboard.html', stats=stats, recent_logs=recent_logs)


@admin_bp.route('/academic_dashboard')
@login_required
@role_required('super_admin', 'academic')
def academic_dashboard():
    """教务仪表板"""
    stats = {
        'class_count': ClassInfo.query.count(),
        'student_count': Student.query.count(),
        'course_count': Course.query.count(),
        'grade_count': Grade.query.count(),
        'fail_count': Grade.query.filter(Grade.score < 60).count(),
        'revision_pending': GradeRevision.query.filter_by(status='pending').count(),
    }
    return render_template('admin/academic_dashboard.html', stats=stats)


# ── 用户管理 ──────────────────────────────────────────
@admin_bp.route('/users')
@login_required
@admin_required
def user_list():
    page = request.args.get('page', 1, type=int)
    role = request.args.get('role', '')
    keyword = request.args.get('keyword', '')
    query = User.query
    if role:
        query = query.filter_by(role=role)
    if keyword:
        query = query.filter(User.username.contains(keyword) | User.name.contains(keyword))
    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('admin/user_list.html', pagination=pagination, role=role, keyword=keyword)


@admin_bp.route('/users/create', methods=['GET', 'POST'])
@login_required
@admin_required
def user_create():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        name = request.form.get('name', '').strip()
        role = request.form.get('role', 'student')
        password = request.form.get('password', '').strip()
        if not username or not name or not password:
            flash('请填写所有必填字段', 'danger')
            return render_template('admin/user_form.html', mode='create')
        if User.query.filter_by(username=username).first():
            flash('用户名已存在', 'danger')
            return render_template('admin/user_form.html', mode='create')
        user = User(username=username, name=name, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()

        # 学生角色自动创建/关联学生信息
        if role == 'student':
            stu = Student.query.filter_by(student_no=username).first()
            if not stu:
                stu = Student(student_no=username, name=name)
                db.session.add(stu)
                db.session.flush()
            if not stu.user_id:
                stu.user_id = user.id
                db.session.commit()
            else:
                flash(f'注意：学号 {username} 已关联到其他账号，未自动关联', 'warning')

        log_operation('create', '创建用户', f'用户 {username} ({role})')
        flash(f'用户 {username} 创建成功', 'success')
        return redirect(url_for('admin.user_list'))
    return render_template('admin/user_form.html', mode='create')


@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def user_edit(user_id):
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.name = request.form.get('name', user.name)
        user.role = request.form.get('role', user.role)
        user.is_active_user = request.form.get('is_active') == 'on'

        # 如果角色改为学生且未关联学生信息，自动创建/关联
        if user.role == 'student' and not user.student:
            stu = Student.query.filter_by(student_no=user.username).first()
            if not stu:
                stu = Student(student_no=user.username, name=user.name)
                db.session.add(stu)
                db.session.flush()
            if not stu.user_id:
                stu.user_id = user.id
                db.session.commit()
            else:
                flash(f'注意：学号 {user.username} 已关联到其他账号，未自动关联', 'warning')


        new_pwd = request.form.get('password', '').strip()
        if new_pwd:
            user.set_password(new_pwd)
        db.session.commit()
        log_operation('update', '编辑用户', f'用户 {user.username}')
        flash('用户信息已更新', 'success')
        return redirect(url_for('admin.user_list'))
    return render_template('admin/user_form.html', mode='edit', user=user)


@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def user_delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.username == 'LYK':
        flash('不能删除超级管理员', 'danger')
        return redirect(url_for('admin.user_list'))
    log_operation('delete', '删除用户', f'用户 {user.username}')
    db.session.delete(user)
    db.session.commit()
    flash('用户已删除', 'success')
    return redirect(url_for('admin.user_list'))


@admin_bp.route('/users/batch_create_students', methods=['GET', 'POST'])
@login_required
@admin_required
def batch_create_students():
    """批量创建学生账号 (基于已有学生信息)"""
    if request.method == 'POST':
        class_id = request.form.get('class_id', type=int)
        query = Student.query
        if class_id:
            query = query.filter_by(class_id=class_id)
        students = query.all()
        created = 0
        for stu in students:
            if not User.query.filter_by(username=stu.student_no).first():
                user = User(username=stu.student_no, name=stu.name, role='student')
                user.set_password(stu.student_no)
                user.student = stu
                db.session.add(user)
                created += 1
        db.session.commit()
        log_operation('create', '批量创建学生账号', f'共创建 {created} 个账号')
        flash(f'成功创建 {created} 个学生账号 (初始密码为学号)', 'success')
        return redirect(url_for('admin.user_list'))
    classes = ClassInfo.query.all()
    return render_template('admin/batch_create.html', classes=classes)


# ── 班级管理 ──────────────────────────────────────────
@admin_bp.route('/classes')
@login_required
@role_required('super_admin', 'academic')
def class_list():
    classes = ClassInfo.query.order_by(ClassInfo.grade.desc()).all()
    return render_template('admin/class_list.html', classes=classes)


@admin_bp.route('/classes/create', methods=['GET', 'POST'])
@admin_bp.route('/classes/<int:class_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('super_admin', 'academic')
def class_form(class_id=None):
    cls = ClassInfo.query.get_or_404(class_id) if class_id else None
    if request.method == 'POST':
        if cls is None:
            cls = ClassInfo()
            db.session.add(cls)
        cls.name = request.form.get('name', '')
        cls.grade = request.form.get('grade', '')
        cls.major = request.form.get('major', '')
        cls.duration = request.form.get('duration', '')
        db.session.commit()
        log_operation('create' if class_id is None else 'update',
                      '班级管理', f'班级 {cls.name}')
        flash('班级信息已保存', 'success')
        return redirect(url_for('admin.class_list'))
    return render_template('admin/class_form.html', cls=cls)


@admin_bp.route('/classes/<int:class_id>/delete', methods=['POST'])
@login_required
@admin_required
def class_delete(class_id):
    cls = ClassInfo.query.get_or_404(class_id)
    if cls.students.count() > 0:
        flash('该班级下还有学生，无法删除', 'danger')
        return redirect(url_for('admin.class_list'))
    log_operation('delete', '删除班级', f'班级 {cls.name}')
    db.session.delete(cls)
    db.session.commit()
    flash('班级已删除', 'success')
    return redirect(url_for('admin.class_list'))


# ── 学生管理 ──────────────────────────────────────────
@admin_bp.route('/students')
@login_required
@role_required('super_admin', 'academic', 'teacher')
def student_list():
    page = request.args.get('page', 1, type=int)
    class_id = request.args.get('class_id', type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', '')
    query = Student.query
    if class_id:
        query = query.filter_by(class_id=class_id)
    if status:
        query = query.filter_by(status=status)
    if keyword:
        query = query.filter(Student.name.contains(keyword) | Student.student_no.contains(keyword))
    pagination = query.order_by(Student.student_no).paginate(page=page, per_page=30, error_out=False)
    classes = ClassInfo.query.all()
    return render_template('admin/student_list.html', pagination=pagination, classes=classes,
                           class_id=class_id, keyword=keyword, status=status)


@admin_bp.route('/students/create', methods=['GET', 'POST'])
@admin_bp.route('/students/<int:student_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('super_admin', 'academic')
def student_form(student_id=None):
    stu = Student.query.get_or_404(student_id) if student_id else None
    classes = ClassInfo.query.all()
    if request.method == 'POST':
        is_new = stu is None
        if is_new:
            stu = Student()
            db.session.add(stu)
        stu.student_no = request.form.get('student_no', '')
        stu.name = request.form.get('name', '')
        stu.gender = request.form.get('gender', '')
        stu.class_id = request.form.get('class_id', type=int)
        stu.college = request.form.get('college', '')
        stu.major = request.form.get('major', '')
        stu.status = request.form.get('status', '正常')
        enroll = request.form.get('enrollment_date', '')
        stu.enrollment_date = datetime.strptime(enroll, '%Y-%m-%d').date() if enroll else None
        db.session.commit()
        # 新建学生时自动创建/关联登录账号（用户名=学号，密码=学号）
        if is_new:
            existing_user = User.query.filter_by(username=stu.student_no).first()
            if not existing_user:
                user = User(username=stu.student_no, name=stu.name, role='student')
                user.set_password(stu.student_no)
                db.session.add(user)
                db.session.flush()  # 先获取 user.id
                stu.user_id = user.id
                db.session.commit()
                flash(f'学生信息已保存，已自动创建账号：{stu.student_no}（密码同学号）', 'success')
            elif not existing_user.student:
                stu.user_id = existing_user.id
                db.session.commit()
                flash(f'学生信息已保存，并已关联到现有账号：{stu.student_no}', 'success')
            else:
                db.session.commit()
                flash(f'学生信息已保存，但账号 {stu.student_no} 已关联其他学生，未自动关联', 'warning')
        log_operation('create' if student_id is None else 'update',
                      '学生管理', f'学生 {stu.name}({stu.student_no})')
        return redirect(url_for('admin.student_list'))
    return render_template('admin/student_form.html', stu=stu, classes=classes)


@admin_bp.route('/students/<int:student_id>/delete', methods=['POST'])
@login_required
@admin_required
def student_delete(student_id):
    stu = Student.query.get_or_404(student_id)
    if stu.grades.count() > 0:
        flash('该学生有成绩记录，无法删除', 'danger')
        return redirect(url_for('admin.student_list'))
    log_operation('delete', '删除学生', f'学生 {stu.name}({stu.student_no})')
    db.session.delete(stu)
    db.session.commit()
    flash('学生已删除', 'success')
    return redirect(url_for('admin.student_list'))


@admin_bp.route('/students/import', methods=['POST'])
@login_required
@role_required('super_admin', 'academic')
def student_import():
    """Excel批量导入学生"""
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('请上传Excel文件 (.xlsx)', 'danger')
        return redirect(url_for('admin.student_list'))
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        imported = 0
        accounts_created = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:
                continue
            try:
                student_no = str(row[0]).strip() if row[0] else ''
                name = str(row[1]).strip() if row[1] else ''
                gender = str(row[2]).strip() if row[2] else ''
                class_name = str(row[3]).strip() if row[3] else ''
                college = str(row[4]).strip() if row[4] else ''
                major = str(row[5]).strip() if row[5] else ''

                if not student_no or not name:
                    continue

                cls = ClassInfo.query.filter_by(name=class_name).first()
                if not cls and class_name:
                    cls = ClassInfo(name=class_name, major=major)
                    db.session.add(cls)
                    db.session.flush()

                existing = Student.query.filter_by(student_no=student_no).first()
                if existing:
                    existing.name = name
                    existing.gender = gender
                    existing.class_id = cls.id if cls else None
                    existing.college = college
                    existing.major = major
                    # 若学生已存在但账号未关联，自动补关联
                    if not existing.user_id:
                        existing_user = User.query.filter_by(username=student_no).first()
                        if existing_user:
                            existing.user_id = existing_user.id
                else:
                    stu = Student(student_no=student_no, name=name, gender=gender,
                                  class_id=cls.id if cls else None, college=college, major=major)
                    db.session.add(stu)
                    db.session.flush()  # 获取学生ID
                    # 自动创建/关联登录账号（用户名=学号，密码=学号）
                    existing_user = User.query.filter_by(username=student_no).first()
                    if not existing_user:
                        user = User(username=student_no, name=name, role='student')
                        user.set_password(student_no)
                        db.session.add(user)
                        db.session.flush()  # 获取 user.id
                        stu.user_id = user.id
                        accounts_created += 1
                    elif not existing_user.student:
                        stu.user_id = existing_user.id
                imported += 1
            except Exception as e:
                errors.append(f'第{row_idx}行: {str(e)}')
        db.session.commit()
        log_operation('import', '导入学生', f'共导入 {imported} 条, 错误 {len(errors)} 条')
        msg = f'成功导入 {imported} 条学生记录'
        if accounts_created > 0:
            msg += f'，自动创建 {accounts_created} 个学生账号（密码同学号）'
        if errors:
            msg += f'，{len(errors)} 条出错'
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'导入失败: {str(e)}', 'danger')
    return redirect(url_for('admin.student_list'))


@admin_bp.route('/students/template')
@login_required
@role_required('super_admin', 'academic')
def student_template():
    """下载导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学生导入模板'
    headers = ['学号', '姓名', '性别', '班级名称', '院系', '专业']
    ws.append(headers)
    ws.append(['423240601', '张三', '男', '23本计算机科学与技术6班', '人工智能学院', '计算机科学与技术'])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='学生导入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 教师管理 ──────────────────────────────────────────
@admin_bp.route('/teachers')
@login_required
@role_required('super_admin', 'academic')
def teacher_list():
    teachers = Teacher.query.order_by(Teacher.teacher_no).all()
    return render_template('admin/teacher_list.html', teachers=teachers)


@admin_bp.route('/teachers/create', methods=['GET', 'POST'])
@admin_bp.route('/teachers/<int:teacher_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('super_admin', 'academic')
def teacher_form(teacher_id=None):
    t = Teacher.query.get_or_404(teacher_id) if teacher_id else None
    if request.method == 'POST':
        if t is None:
            t = Teacher()
            db.session.add(t)
        t.teacher_no = request.form.get('teacher_no', '')
        t.name = request.form.get('name', '')
        t.department = request.form.get('department', '')
        db.session.commit()
        log_operation('create' if teacher_id is None else 'update',
                      '教师管理', f'教师 {t.name}({t.teacher_no})')
        flash('教师信息已保存', 'success')
        return redirect(url_for('admin.teacher_list'))
    return render_template('admin/teacher_form.html', t=t)


@admin_bp.route('/teachers/<int:teacher_id>/delete', methods=['POST'])
@login_required
@admin_required
def teacher_delete(teacher_id):
    t = Teacher.query.get_or_404(teacher_id)
    if t.courses.count() > 0:
        flash('该教师有授课记录，无法删除', 'danger')
        return redirect(url_for('admin.teacher_list'))
    db.session.delete(t)
    db.session.commit()
    flash('教师已删除', 'success')
    return redirect(url_for('admin.teacher_list'))


# ── 课程管理 ──────────────────────────────────────────
@admin_bp.route('/courses')
@login_required
@role_required('super_admin', 'academic', 'teacher')
def course_list():
    courses = Course.query.order_by(Course.semester.desc(), Course.name).all()
    return render_template('admin/course_list.html', courses=courses)


@admin_bp.route('/courses/create', methods=['GET', 'POST'])
@admin_bp.route('/courses/<int:course_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('super_admin', 'academic')
def course_form(course_id=None):
    c = Course.query.get_or_404(course_id) if course_id else None
    teachers = Teacher.query.all()
    classes = ClassInfo.query.all()
    if request.method == 'POST':
        if c is None:
            c = Course()
            db.session.add(c)
        c.course_code = request.form.get('course_code', '')
        c.name = request.form.get('name', '')
        c.credit = request.form.get('credit', 2.0, type=float)
        c.hours = request.form.get('hours', 32, type=int)
        c.assessment_type = request.form.get('assessment_type', '考查')
        c.teacher_id = request.form.get('teacher_id', type=int)
        c.class_id = request.form.get('class_id', type=int)
        c.semester = request.form.get('semester', '2025-2026-2')
        db.session.commit()
        log_operation('create' if course_id is None else 'update',
                      '课程管理', f'课程 {c.name}({c.course_code})')
        flash('课程信息已保存', 'success')
        return redirect(url_for('admin.course_list'))
    return render_template('admin/course_form.html', c=c, teachers=teachers, classes=classes)


@admin_bp.route('/courses/<int:course_id>/delete', methods=['POST'])
@login_required
@admin_required
def course_delete(course_id):
    c = Course.query.get_or_404(course_id)
    if c.grades.count() > 0:
        flash('该课程有成绩记录，无法删除', 'danger')
        return redirect(url_for('admin.course_list'))
    db.session.delete(c)
    db.session.commit()
    flash('课程已删除', 'success')
    return redirect(url_for('admin.course_list'))


@admin_bp.route('/courses/<int:course_id>/lock', methods=['POST'])
@login_required
@role_required('super_admin', 'academic')
def course_lock(course_id):
    c = Course.query.get_or_404(course_id)
    c.is_locked = not c.is_locked
    db.session.commit()
    log_operation('update', '课程锁定', f'课程 {c.name} {"锁定" if c.is_locked else "解锁"}')
    flash(f'课程已{"锁定" if c.is_locked else "解锁"}', 'success')
    return redirect(url_for('admin.course_list'))


# ── 成绩修改审核 ──────────────────────────────────────────
@admin_bp.route('/revisions')
@login_required
@role_required('super_admin', 'academic')
def revision_list():
    status = request.args.get('status', 'pending')
    query = GradeRevision.query
    if status:
        query = query.filter_by(status=status)
    revisions = query.order_by(GradeRevision.created_at.desc()).all()
    return render_template('admin/revision_list.html', revisions=revisions, status=status)


@admin_bp.route('/revisions/<int:rev_id>/approve', methods=['POST'])
@login_required
@role_required('super_admin', 'academic')
def revision_approve(rev_id):
    rev = GradeRevision.query.get_or_404(rev_id)
    if rev.status != 'pending':
        flash('该申请已处理', 'warning')
        return redirect(url_for('admin.revision_list'))
    rev.status = 'approved'
    rev.reviewed_by = current_user.id
    rev.reviewed_at = datetime.now()
    grade = rev.grade
    grade.score = rev.new_score
    grade.special_mark = rev.new_special or None
    grade.gpa = Grade.calc_gpa(rev.new_score, rev.new_special)
    db.session.commit()
    log_operation('update', '审核成绩修改', f'已批准: {grade.student.name} {grade.course.name}')
    flash('已批准成绩修改', 'success')
    return redirect(url_for('admin.revision_list'))


@admin_bp.route('/revisions/<int:rev_id>/reject', methods=['POST'])
@login_required
@role_required('super_admin', 'academic')
def revision_reject(rev_id):
    rev = GradeRevision.query.get_or_404(rev_id)
    rev.status = 'rejected'
    rev.reviewed_by = current_user.id
    rev.reviewed_at = datetime.now()
    rev.detail = request.form.get('reason', '')
    db.session.commit()
    flash('已驳回成绩修改申请', 'info')
    return redirect(url_for('admin.revision_list'))


# ── 日志管理 ──────────────────────────────────────────
@admin_bp.route('/logs/operation')
@login_required
@role_required('super_admin', 'academic')
def operation_logs():
    page = request.args.get('page', 1, type=int)
    keyword = request.args.get('keyword', '')
    query = OperationLog.query
    if keyword:
        query = query.filter(OperationLog.target.contains(keyword) | OperationLog.action.contains(keyword))
    pagination = query.order_by(OperationLog.created_at.desc()).paginate(page=page, per_page=30, error_out=False)
    return render_template('admin/operation_logs.html', pagination=pagination, keyword=keyword)


@admin_bp.route('/logs/login')
@login_required
@admin_required
def login_logs():
    page = request.args.get('page', 1, type=int)
    pagination = LoginLog.query.order_by(LoginLog.created_at.desc()).paginate(page=page, per_page=30, error_out=False)
    return render_template('admin/login_logs.html', pagination=pagination)


# ── 通知公告 ──────────────────────────────────────────
@admin_bp.route('/notifications')
@login_required
@role_required('super_admin', 'academic')
def notification_list():
    notifs = Notification.query.order_by(Notification.created_at.desc()).all()
    return render_template('admin/notification_list.html', notifs=notifs)


@admin_bp.route('/notifications/create', methods=['GET', 'POST'])
@admin_bp.route('/notifications/<int:nid>/edit', methods=['GET', 'POST'])
@login_required
@role_required('super_admin', 'academic')
def notification_form(nid=None):
    n = Notification.query.get_or_404(nid) if nid else None
    if request.method == 'POST':
        if n is None:
            n = Notification(created_by=current_user.id)
            db.session.add(n)
        n.title = request.form.get('title', '')
        n.content = request.form.get('content', '')
        n.target_role = request.form.get('target_role', 'all')
        n.is_pinned = request.form.get('is_pinned') == 'on'
        db.session.commit()
        log_operation('create' if nid is None else 'update', '通知公告', f'通知: {n.title}')
        flash('通知已发布', 'success')
        return redirect(url_for('admin.notification_list'))
    return render_template('admin/notification_form.html', n=n)


@admin_bp.route('/notifications/<int:nid>/delete', methods=['POST'])
@login_required
@admin_required
def notification_delete(nid):
    n = Notification.query.get_or_404(nid)
    db.session.delete(n)
    db.session.commit()
    flash('通知已删除', 'success')
    return redirect(url_for('admin.notification_list'))


# ── 系统设置 ──────────────────────────────────────────
@admin_bp.route('/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    if request.method == 'POST':
        for key in ['max_score', 'pass_score', 'excellent_score', 'current_semester', 'gpa_rule']:
            val = request.form.get(key, '')
            setting = SystemSetting.query.filter_by(key=key).first()
            if setting:
                setting.value = val
            else:
                db.session.add(SystemSetting(key=key, value=val, description=key))
        db.session.commit()
        log_operation('update', '系统设置', '修改系统参数')
        flash('设置已保存', 'success')
        return redirect(url_for('admin.settings'))
    settings_dict = {s.key: s.value for s in SystemSetting.query.all()}
    return render_template('admin/settings.html', settings=settings_dict)


# ── 数据备份 ──────────────────────────────────────────
@admin_bp.route('/backup', methods=['GET', 'POST'])
@login_required
@admin_required
def backup():
    if request.method == 'POST':
        backup_dir = current_app.config['BACKUP_FOLDER']
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f'backup_{timestamp}'
        backup_path = os.path.join(backup_dir, backup_name)
        db_path = current_app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
        if os.path.exists(db_path):
            shutil.copy2(db_path, os.path.join(backup_dir, f'grade_system_{timestamp}.db'))
            log_operation('export', '数据备份', f'备份至 {backup_name}')
            flash(f'备份成功: grade_system_{timestamp}.db', 'success')
        else:
            flash('数据库文件不存在', 'danger')
        return redirect(url_for('admin.backup'))
    backup_dir = current_app.config['BACKUP_FOLDER']
    backups = []
    if os.path.exists(backup_dir):
        for f in sorted(os.listdir(backup_dir), reverse=True):
            if f.endswith('.db'):
                fpath = os.path.join(backup_dir, f)
                backups.append({
                    'name': f,
                    'size': f'{os.path.getsize(fpath) / 1024:.1f} KB',
                    'time': datetime.fromtimestamp(os.path.getmtime(fpath)).strftime('%Y-%m-%d %H:%M:%S')
                })
    return render_template('admin/backup.html', backups=backups)


@admin_bp.route('/backup/<filename>')
@login_required
@admin_required
def backup_download(filename):
    backup_dir = current_app.config['BACKUP_FOLDER']
    return send_file(os.path.join(backup_dir, filename), as_attachment=True)
