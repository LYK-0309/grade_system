"""
工具函数: 装饰器、日志、Excel处理、统计计算
"""
from functools import wraps
from flask import flash, redirect, url_for, abort, request, jsonify
from flask_login import current_user
from app.models import db, LoginLog, OperationLog
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── 权限装饰器 ──────────────────────────────────
def role_required(*roles):
    """角色权限检查"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('请先登录', 'warning')
                return redirect(url_for('auth.login'))
            if current_user.role not in roles:
                abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def admin_required(f):
    return role_required('super_admin')(f)


def teacher_or_admin(f):
    return role_required('super_admin', 'teacher', 'academic')(f)


def academic_or_admin(f):
    return role_required('super_admin', 'academic')(f)


# ── 日志记录 ──────────────────────────────────
def log_login(user, action, detail=''):
    log = LoginLog(
        user_id=user.id if user else None,
        username=user.username if user else request.form.get('username', ''),
        ip=request.remote_addr or '127.0.0.1',
        action=action,
        detail=detail
    )
    db.session.add(log)
    db.session.commit()


def log_operation(action, target, detail=''):
    log = OperationLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        username=current_user.username if current_user.is_authenticated else 'system',
        action=action,
        target=target,
        detail=detail
    )
    db.session.add(log)
    db.session.commit()


# ── Excel导出工具 ──────────────────────────────────
def create_grade_export(workbook_title, headers, rows, summary_row=None):
    """创建成绩导出Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = workbook_title[:31]

    # 标题行
    title_font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill('solid', start_color='2563EB')
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='3B82F6')
    data_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1, value=workbook_title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[2].height = 25

    # 数据
    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # 汇总行
    if summary_row:
        row_idx = len(rows) + 3
        for col_idx, value in enumerate(summary_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Arial', bold=True, size=10)
            cell.fill = PatternFill('solid', start_color='DBEAFE')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # 自动列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(3, len(rows) + 3):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 30)

    return wb


def export_to_bytes(wb):
    """Workbook转bytes"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 统计计算 ──────────────────────────────────
def calc_class_stats(grades):
    """计算班级统计"""
    valid_grades = [g for g in grades if g.special_mark not in ('缺考', '作弊') and g.score is not None]
    scores = [g.score for g in valid_grades]

    if not scores:
        return {
            'count': len(grades), 'valid_count': 0, 'avg': 0, 'max': 0, 'min': 0,
            'pass_count': 0, 'pass_rate': 0, 'fail_count': 0, 'excellent_rate': 0,
            'std': 0, 'gpa_avg': 0
        }

    pass_count = sum(1 for s in scores if s >= 60)
    excellent_count = sum(1 for s in scores if s >= 90)
    avg = sum(scores) / len(scores)
    variance = sum((s - avg) ** 2 for s in scores) / len(scores) if scores else 0

    gpa_list = [g.gpa or 0 for g in valid_grades if g.gpa is not None]

    return {
        'count': len(grades),
        'valid_count': len(scores),
        'avg': round(avg, 2),
        'max': round(max(scores), 2),
        'min': round(min(scores), 2),
        'pass_count': pass_count,
        'pass_rate': round(pass_count / len(scores) * 100, 1),
        'fail_count': len(scores) - pass_count,
        'excellent_rate': round(excellent_count / len(scores) * 100, 1),
        'std': round(variance ** 0.5, 2),
        'gpa_avg': round(sum(gpa_list) / len(gpa_list), 2) if gpa_list else 0,
    }


def calc_grade_distribution(grades):
    """成绩分布 (用于柱状图)"""
    ranges = {'90-100': 0, '80-89': 0, '70-79': 0, '60-69': 0, '0-59': 0, '缺考/作弊': 0}
    for g in grades:
        if g.special_mark in ('缺考', '作弊'):
            ranges['缺考/作弊'] += 1
        elif g.score is None:
            ranges['缺考/作弊'] += 1
        elif g.score >= 90:
            ranges['90-100'] += 1
        elif g.score >= 80:
            ranges['80-89'] += 1
        elif g.score >= 70:
            ranges['70-79'] += 1
        elif g.score >= 60:
            ranges['60-69'] += 1
        else:
            ranges['0-59'] += 1
    return ranges


def get_rankings(grades, by='score'):
    """计算排名"""
    valid = [g for g in grades if g.special_mark not in ('缺考', '作弊') and g.score is not None]
    valid.sort(key=lambda g: g.score, reverse=True)
    rankings = {}
    for rank, g in enumerate(valid, 1):
        rankings[g.student_id] = rank
    return rankings
